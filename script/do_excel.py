from openpyxl import load_workbook
from script.path_split import NEW_API_DATA_PATH,ORDER_API_DATA_PATH


class HandleExcel:
    '''
    封装读取excle用例类
    '''

    def __init__(self,filename,sheet=None):
        self.filename = filename
        self.sheet = sheet
        wb = load_workbook(self.filename)
        # 判断是否传入sheet值
        if sheet:
            self.ws = wb[self.sheet]
        else:
            self.ws = wb.active # 没有传入默认取第一个sheet
        self.lists = []
        self.first_rows = tuple(self.ws.iter_rows(max_row=1,values_only=True))[0] # 获取首行数据

    def get_all_datas(self):
        '''
        获取sheet中所有数据组成的嵌套字典的列表
        :return: 返回列表值
        '''
        for rows in self.ws.iter_rows(min_row=2,values_only=True):
            self.lists.append(dict(zip(self.first_rows,rows)))
        return self.lists

    def get_one_datas(self,row_num):
        '''
        返回某行值
        :param row_num: 行号
        :return: 列表值
        '''
        one_datas = []
        if isinstance(row_num, int):
            if 1 < row_num <= self.ws.max_row:

                one_datas = tuple(self.ws.iter_rows(min_row=row_num,max_row=row_num,values_only=True))[0]
            else:
                print('row invalid')
            return one_datas

    def write_datas(self,other_file,other_sheet,write_num,actual_result,result):
        '''
        往sheet中写入数据
        :param other_file: 文件名称
        :param other_sheet: 表格名称
        :param write_num: 写入行号
        :param actual_result: 写入实际结果
        :param result: 测试用例执行结果
        :return:
        '''
        wb = load_workbook(other_file)
        ws = wb[other_sheet]
        if isinstance(write_num,int)and 1<write_num<=ws.max_row:
            ws.cell(row=write_num,column=8,value=actual_result)
            ws.cell(row=write_num,column=9,value=result)
        wb.save(other_file) # 注意excel中写入数据需要保存单元格


handle_new_api_reg = HandleExcel(NEW_API_DATA_PATH, sheet='reg')
handle_new_api_recharge = HandleExcel(NEW_API_DATA_PATH, sheet='recharge')
handle_new_api_invest = HandleExcel(NEW_API_DATA_PATH, sheet='invest')
handle_order_login = HandleExcel(ORDER_API_DATA_PATH,sheet='login')
handle_order_addsddress = HandleExcel(ORDER_API_DATA_PATH, sheet='add_address')


if __name__ == '__main__':
    '''#print(handle_aidong_excel.get_all_datas())
    #print(handle_excel.get_one_datas(4))
    # print(handle_excel.write_datas(NEW_DATA_PATH_FILE,'home',2,'登录成功','ok'))
    handel_excel = HandleExcel(AIDONG_DATA_PATH_FILE, sheet='coursedetails')
    case = handel_excel.get_all_datas()
    print(case)'''
    # print(handle_new_api_invest.get_all_datas())
    print(handle_order_login.get_all_datas())