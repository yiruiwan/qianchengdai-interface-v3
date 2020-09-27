from openpyxl import load_workbook
from script.rebulit_config import do_config
from script.path_split import DATA_PATH_FILE


class ReadExcel:
    def __init__(self, sheet=None):
        '''
        :param sheet: 表单名
        '''
        self.filename = DATA_PATH_FILE
        self.sheet = sheet
        # 打开文件
        self.wb = load_workbook(self.filename)
        # 定位表单
        if self.sheet is None:
            self.ws = self.wb.active  # 定位到第一个表单
        else:
            self.ws = self.wb[self.sheet]  # 根据表单名定位
        self.case = []
        # 获取表头
        self.first_row = tuple(self.ws.iter_rows(max_row=1, values_only=True))[0]

    def get_case(self):
        '''
        获取全部Excel数据
        :return: 返回嵌套字典的列表
        '''
        for rows in self.ws.iter_rows(min_row=2, values_only=True):
            self.case.append(dict(zip(self.first_row, rows)))
        return self.case

    def one_case(self, row):
        '''
        获取某一行Excel数据
        :param row: 需要返回的行号
        '''
        # 判断输入的行号是否超出范围
        if isinstance(row, int) and 1 < row <= self.ws.max_row:
            case = tuple(self.ws.iter_rows(min_row=row, max_row=row, values_only=True))[0]
            print(case)
        else:
            print('行数输入不对')

    def write_case(self, row_number, finally_result, result):
        '''
        Excel中写入数据
        :param row_number: 写入的行号
        :param finally_result: 预期结果值
        :param result: 测试结果
        '''
        other_wb = load_workbook(self.filename)  # 定位写入哪个文件
        other_ws = other_wb[self.sheet]  # 定位写入哪个表单
        if isinstance(row_number, int) and 1 < row_number <= self.ws.max_row:
            other_ws.cell(row=row_number, column=do_config('excel', 'one_column'), value=finally_result)
            other_ws.cell(row=row_number, column=do_config('excel', 'two_column'), value=result)
            other_wb.save(self.filename)
        else:
            pass


if __name__ == '__main__':
    readexcel = ReadExcel(sheet='recharge')
    print(readexcel.get_case())
    # print(readexcel.one_case(2))
    # print(readexcel.write_case(4, 10, 'ok'))
